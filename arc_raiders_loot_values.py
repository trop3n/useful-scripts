#!/usr/bin/env python3
"""
Arc Raiders Loot Value Database Generator
Creates a spreadsheet with all lootable items organized by tier and value.
Data sourced from MetaForge community database (metaforge.app/arc-raiders)

Requirements: pip install openpyxl
Usage: python arc_raiders_loot_values.py
Output: arc_raiders_loot_values.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Loot items organized by tier
# Format: (Item Name, Sell Value, Category)

S_TIER = [  # 5,000+ coins - Legendary/Epic drops
    ("Snap Hook", 14000, "Utility"),
    ("Queen Reactor", 13000, "ARC Part"),
    ("Matriarch Reactor", 13000, "ARC Part"),
    ("Lance's Mixtape 5th Ed", 10000, "Collectible"),
    ("Horizontal Grip", 7000, "Attachment"),
    ("Magnetic Accelerator", 5500, "Refined Material"),
    ("Power Rod", 5500, "Refined Material"),
    ("Wolfpack", 5000, "Grenade"),
    ("Deadline", 5000, "Grenade"),
    ("Photoelectric Cloak", 5000, "Utility"),
    ("Jupiter", 5000, "Weapon"),
    ("Equalizer", 5000, "Weapon"),
    ("Aphelion", 5000, "Weapon"),
    ("Bobcat I", 5000, "Weapon"),
    ("Vulcano I", 5000, "Weapon"),
    ("Hullcracker I", 5000, "Weapon"),
    ("Bettina I", 5000, "Weapon"),
    ("Tempest I", 5000, "Weapon"),
    ("Blueprint (Any)", 5000, "Blueprint"),
    ("Angled Grip III", 5000, "Attachment"),
    ("Vertical Grip III", 5000, "Attachment"),
    ("Stable Stock III", 5000, "Attachment"),
    ("Compensator III", 5000, "Attachment"),
    ("Muzzle Brake III", 5000, "Attachment"),
    ("Shotgun Choke III", 5000, "Attachment"),
    ("Extended Light Mag III", 5000, "Attachment"),
    ("Extended Medium Mag III", 5000, "Attachment"),
    ("Extended Barrel", 5000, "Attachment"),
    ("Shotgun Silencer", 5000, "Attachment"),
    ("Lightweight Stock", 5000, "Attachment"),
]

A_TIER = [  # 3,000-3,500 coins - Rare drops
    ("Osprey I", 3500, "Weapon"),
    ("Torrente I", 3500, "Weapon"),
    ("Venator I", 3500, "Weapon"),
    ("Renegade I", 3500, "Weapon"),
    ("Complex Gun Parts", 3000, "Material"),
    ("Exodus Modules", 2750, "ARC Part"),
]

B_TIER = [  # 1,000-2,880 coins - Uncommon valuable
    ("Raider Hatch Key", 2000, "Key"),
    ("Anvil I", 2000, "Weapon"),
    ("Il Toro I", 2000, "Weapon"),
    ("Arpeggio I", 2000, "Weapon"),
    ("Angled Grip II", 2000, "Attachment"),
    ("Vertical Grip II", 2000, "Attachment"),
    ("Stable Stock II", 2000, "Attachment"),
    ("Compensator II", 2000, "Attachment"),
    ("Muzzle Brake II", 2000, "Attachment"),
    ("Shotgun Choke II", 2000, "Attachment"),
    ("Extended Light Mag II", 2000, "Attachment"),
    ("Extended Medium Mag II", 2000, "Attachment"),
    ("Advanced Mechanical Components", 1750, "Refined Material"),
    ("Advanced Electrical Components", 1750, "Refined Material"),
    ("Mod Components", 1750, "Refined Material"),
    ("Heavy Fuze Grenade", 1600, "Grenade"),
    ("Blaze Grenade", 1600, "Grenade"),
    ("Burletta I", 1500, "Weapon"),
    ("ARC Circuitry", 1000, "ARC Part"),
    ("ARC Motion Core", 1000, "ARC Part"),
    ("Explosive Compound", 1000, "Material"),
    ("Antiseptic", 1000, "Medical"),
    ("Industrial Magnet", 1000, "Material"),
    ("Smoke Grenade", 1000, "Grenade"),
    ("Lure Grenade", 1000, "Grenade"),
    ("Tagging Grenade", 1000, "Grenade"),
    ("Zipline", 1000, "Utility"),
    ("Trigger 'Nade", 1000, "Grenade"),
]

C_TIER = [  # 500-850 coins - Common valuable
    ("Jolt Mine", 850, "Trap"),
    ("Shrapnel Grenade", 800, "Grenade"),
    ("Snap Blast Grenade", 800, "Grenade"),
    ("Light Gun Parts", 700, "Material"),
    ("Medium Gun Parts", 700, "Material"),
    ("Heavy Gun Parts", 700, "Material"),
    ("Synthesized Fuel", 700, "Material"),
    ("Mechanical Components", 640, "Refined Material"),
    ("Electrical Components", 640, "Refined Material"),
    ("Durable Cloth", 640, "Refined Material"),
    ("Angled Grip I", 640, "Attachment"),
    ("Vertical Grip I", 640, "Attachment"),
    ("Stable Stock I", 640, "Attachment"),
    ("Compensator I", 640, "Attachment"),
    ("Muzzle Brake I", 640, "Attachment"),
    ("Shotgun Choke I", 640, "Attachment"),
    ("Extended Light Mag I", 640, "Attachment"),
    ("Extended Medium Mag I", 640, "Attachment"),
    ("Extended Shotgun Mag I", 640, "Attachment"),
    ("Binoculars", 640, "Utility"),
    ("Green Light Stick", 640, "Utility"),
    ("Remote Raider Flare", 640, "Utility"),
    ("Barricade Kit", 640, "Utility"),
    ("Rope", 500, "Material"),
    ("Speaker Component", 500, "Material"),
    ("Sensors", 500, "Material"),
    ("Processor", 500, "Material"),
    ("Voltage Converter", 500, "Material"),
    ("Hairpin I", 500, "Weapon"),
    ("Ferro I", 500, "Weapon"),
    ("Kettle I", 500, "Weapon"),
    ("Rattler I", 500, "Weapon"),
    ("Stitcher I", 500, "Weapon"),
]

D_TIER = [  # 50-470 coins - Basic materials
    ("Simple Gun Parts", 330, "Material"),
    ("Duct Tape", 300, "Material"),
    ("Steel Spring", 300, "Material"),
    ("Canister", 300, "Material"),
    ("Oil", 300, "Material"),
    ("Gas Grenade", 300, "Grenade"),
    ("Light Impact Grenade", 300, "Grenade"),
    ("Li'l Smoke Grenade", 300, "Utility"),
    ("Door Blocker", 300, "Utility"),
    ("Great Mullein", 300, "Plant"),
    ("Crude Explosives", 270, "Material"),
    ("Power Cell", 270, "ARC Part"),
    ("ARC Powercell", 250, "ARC Part"),
    ("ARC Alloy", 200, "ARC Part"),
    ("Battery", 200, "Material"),
    ("Wires", 100, "Material"),
    ("Metal Parts", 50, "Material"),
    ("Plastic Parts", 50, "Material"),
    ("Rubber Parts", 50, "Material"),
    ("Fabric", 50, "Material"),
    ("Chemicals", 50, "Material"),
]

def create_loot_spreadsheet(output_path="arc_raiders_loot_values.xlsx"):
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    tier_font = Font(bold=True, size=12)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    currency_format = '"$"#,##0'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Tier colors
    tier_colors = {
        "S": PatternFill("solid", fgColor="FFD700"),  # Gold
        "A": PatternFill("solid", fgColor="C0C0C0"),  # Silver
        "B": PatternFill("solid", fgColor="CD7F32"),  # Bronze
        "C": PatternFill("solid", fgColor="90EE90"),  # Light Green
        "D": PatternFill("solid", fgColor="D3D3D3"),  # Light Gray
    }
    
    header_fill = PatternFill("solid", fgColor="1565C0")
    
    # Sheet 1: All Items by Tier
    ws_all = wb.active
    ws_all.title = "All Loot by Tier"
    
    headers = ["Item Name", "Sell Value", "Tier", "Category"]
    for col, header in enumerate(headers, 1):
        cell = ws_all.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    all_items = []
    for tier_name, tier_data in [("S", S_TIER), ("A", A_TIER), ("B", B_TIER), ("C", C_TIER), ("D", D_TIER)]:
        for item_name, value, category in tier_data:
            all_items.append((item_name, value, tier_name, category))
    
    # Sort by value descending
    all_items.sort(key=lambda x: -x[1])
    
    row = 2
    for item_name, value, tier, category in all_items:
        ws_all.cell(row=row, column=1, value=item_name).alignment = left_align
        ws_all.cell(row=row, column=2, value=value).number_format = currency_format
        ws_all.cell(row=row, column=3, value=tier).alignment = center_align
        ws_all.cell(row=row, column=4, value=category).alignment = center_align
        
        for col in range(1, 5):
            ws_all.cell(row=row, column=col).border = thin_border
            ws_all.cell(row=row, column=col).fill = tier_colors[tier]
        row += 1
    
    ws_all.column_dimensions['A'].width = 32
    ws_all.column_dimensions['B'].width = 12
    ws_all.column_dimensions['C'].width = 8
    ws_all.column_dimensions['D'].width = 18
    
    # Sheet 2: Summary Statistics
    ws_summary = wb.create_sheet("Summary")
    ws_summary['A1'] = "Arc Raiders Loot Value Database"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary['A2'] = "Data Source: MetaForge Community Database (Dec 2025)"
    
    ws_summary['A4'] = "Tier Breakdown:"
    ws_summary['A4'].font = Font(bold=True)
    
    tier_stats = [
        ("S-Tier (Legendary)", "$5,000 - $14,000", len(S_TIER)),
        ("A-Tier (Epic)", "$2,750 - $3,500", len(A_TIER)),
        ("B-Tier (Rare)", "$1,000 - $2,000", len(B_TIER)),
        ("C-Tier (Uncommon)", "$500 - $850", len(C_TIER)),
        ("D-Tier (Common)", "$50 - $470", len(D_TIER)),
    ]
    
    for i, (tier_name, value_range, count) in enumerate(tier_stats, 5):
        ws_summary[f'A{i}'] = tier_name
        ws_summary[f'B{i}'] = value_range
        ws_summary[f'C{i}'] = f"{count} items"
    
    ws_summary['A11'] = f"Total Items: {len(all_items)}"
    ws_summary['A11'].font = Font(bold=True)
    
    ws_summary['A13'] = "Top 10 Most Valuable Items:"
    ws_summary['A13'].font = Font(bold=True)
    
    top_items = sorted(all_items, key=lambda x: -x[1])[:10]
    for i, (name, value, tier, cat) in enumerate(top_items, 14):
        ws_summary[f'A{i}'] = f"{name}: ${value:,} ({tier}-Tier)"
    
    ws_summary['A25'] = "Expedition Strategy Notes:"
    ws_summary['A25'].font = Font(bold=True)
    ws_summary['A26'] = "• Stash value = items + cash combined"
    ws_summary['A27'] = "• 1 million coins = 1 skill point (max 5 at 5 million)"
    ws_summary['A28'] = "• Prioritize S-tier ARC parts (Queen/Matriarch Reactor)"
    ws_summary['A29'] = "• Blueprints always sell for $5,000"
    ws_summary['A30'] = "• Snap Hook is the highest value item at $14,000"
    
    ws_summary.column_dimensions['A'].width = 45
    ws_summary.column_dimensions['B'].width = 18
    ws_summary.column_dimensions['C'].width = 12
    
    # Sheet 3: By Category
    ws_cat = wb.create_sheet("By Category")
    
    categories = {}
    for item_name, value, tier, category in all_items:
        if category not in categories:
            categories[category] = []
        categories[category].append((item_name, value, tier))
    
    for col, header in enumerate(["Item Name", "Sell Value", "Tier"], 1):
        cell = ws_cat.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    row = 2
    for category in sorted(categories.keys()):
        # Category header
        ws_cat.cell(row=row, column=1, value=f"=== {category.upper()} ===")
        ws_cat.cell(row=row, column=1).font = tier_font
        row += 1
        
        # Items sorted by value
        for item_name, value, tier in sorted(categories[category], key=lambda x: -x[1]):
            ws_cat.cell(row=row, column=1, value=item_name).alignment = left_align
            ws_cat.cell(row=row, column=2, value=value).number_format = currency_format
            ws_cat.cell(row=row, column=3, value=tier).alignment = center_align
            for col in range(1, 4):
                ws_cat.cell(row=row, column=col).border = thin_border
                ws_cat.cell(row=row, column=col).fill = tier_colors[tier]
            row += 1
        row += 1  # Blank row between categories
    
    ws_cat.column_dimensions['A'].width = 32
    ws_cat.column_dimensions['B'].width = 12
    ws_cat.column_dimensions['C'].width = 8
    
    wb.save(output_path)
    print(f"Created {output_path} with {len(all_items)} items")
    print(f"  S-Tier: {len(S_TIER)} items")
    print(f"  A-Tier: {len(A_TIER)} items")
    print(f"  B-Tier: {len(B_TIER)} items")
    print(f"  C-Tier: {len(C_TIER)} items")
    print(f"  D-Tier: {len(D_TIER)} items")

if __name__ == "__main__":
    create_loot_spreadsheet()
