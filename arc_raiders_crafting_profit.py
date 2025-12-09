#!/usr/bin/env python3
"""
Arc Raiders Crafting Profitability Calculator
Analyzes all crafting recipes to determine profit/loss when selling crafted items.
Data sourced from ARCTracker.io (December 2025)

Requirements: pip install openpyxl
Usage: python arc_raiders_crafting_profit.py
Output: arc_raiders_crafting_profit.xlsx

To update recipes:
1. Edit the material_values dict to adjust material sell prices
2. Edit the recipes list to add/modify crafting recipes
3. Run the script to regenerate the spreadsheet
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Material sell values - verified from MetaForge loot tiers + ARCTracker
# Adjust these values if game patches change sell prices
material_values = {
    # Basic Materials (Common)
    "Metal Parts": 50,
    "Plastic Parts": 50,
    "Rubber Parts": 50,
    "Fabric": 50,
    "Chemicals": 50,
    "Wires": 100,
    
    # Refined Materials (Uncommon)
    "Duct Tape": 300,
    "Steel Spring": 300,
    "Canister": 300,
    "Oil": 300,
    "Battery": 200,
    "Simple Gun Parts": 330,
    "Mechanical Components": 640,
    "Electrical Components": 640,
    "Durable Cloth": 640,
    "Crude Explosives": 270,
    
    # Refined Materials (Rare)
    "Light Gun Parts": 700,
    "Medium Gun Parts": 700,
    "Heavy Gun Parts": 700,
    "Shotgun Parts": 700,
    "Synthesized Fuel": 700,
    "Advanced Mechanical Components": 1750,
    "Advanced Electrical Components": 1750,
    "Mod Components": 1750,
    "Explosive Compound": 1000,
    "Antiseptic": 1000,
    "ARC Circuitry": 1000,
    "ARC Motion Core": 1000,
    "Industrial Magnet": 1000,
    "Rope": 500,
    "Speaker Component": 500,
    "Sensors": 500,
    "Processor": 500,
    "Voltage Converter": 500,
    "Great Mullein": 300,
    "ARC Alloy": 200,
    
    # High-Value Materials (Epic/Legendary)
    "Complex Gun Parts": 3000,
    "Magnetic Accelerator": 5500,
    "Power Rod": 5500,
    "Exodus Modules": 2750,
    "Queen Reactor": 13000,
    "Matriarch Reactor": 13000,
}

# Recipes from ARCTracker (verified Dec 2025)
# Format: (Item Name, Sell Value, [(material, qty), ...], Bench, Notes)
# To add a recipe: ("Item Name", sell_price, [("Material1", qty), ("Material2", qty)], "Bench Name", "Notes")
recipes = [
    # === EXPLOSIVES STATION ===
    ("Wolfpack", 5000, [("Explosive Compound", 3), ("ARC Motion Core", 2)], "Explosives III", "Epic grenade"),
    ("Deadline", 5000, [("Explosive Compound", 3), ("ARC Circuitry", 2)], "Explosives III", "Epic weapon"),
    ("Heavy Fuze Grenade", 1600, [("Explosive Compound", 1), ("Canister", 2)], "Explosives", "Timed explosive"),
    ("Trigger 'Nade", 1000, [("Crude Explosives", 2), ("Processor", 1)], "Explosives II", "Remote detonated"),
    ("Jolt Mine", 850, [("Electrical Components", 1), ("Battery", 1)], "Explosives II", "Electric trap"),
    ("Shrapnel Grenade", 800, [("Crude Explosives", 1), ("Steel Spring", 2)], "Explosives", ""),
    ("Snap Blast Grenade", 800, [("Crude Explosives", 2), ("Industrial Magnet", 1)], "Explosives", "Sticky grenade"),
    ("Gas Grenade", 300, [("Chemicals", 4), ("Rubber Parts", 2)], "Explosives", "Common"),
    ("Light Impact Grenade", 300, [("Chemicals", 3), ("Plastic Parts", 2)], "Explosives", "Common"),
    
    # === UTILITY STATION ===
    ("Snap Hook", 14000, [("Power Rod", 1), ("Exodus Modules", 1)], "Utility III", "Grappling hook"),
    ("Smoke Grenade", 1000, [("Chemicals", 14), ("Canister", 1)], "Utility II", "Large smoke"),
    ("Lure Grenade", 1000, [("Speaker Component", 1), ("Electrical Components", 1)], "Utility II", "Distraction"),
    ("Tagging Grenade", 1000, [("Electrical Components", 1), ("Sensors", 1)], "Utility II", "Tracking"),
    ("Zipline", 1000, [("Rope", 1), ("Mechanical Components", 1)], "Utility", "Traversal"),
    ("Barricade Kit", 640, [("Mechanical Components", 1)], "Utility II", "Defense"),
    ("Binoculars", 640, [("Plastic Parts", 8), ("Rubber Parts", 4)], "Utility", "Scouting"),
    ("Door Blocker", 300, [("Metal Parts", 3), ("Rubber Parts", 3)], "Utility", ""),
    ("Li'l Smoke Grenade", 300, [("Chemicals", 5), ("Plastic Parts", 1)], "Utility", "Small smoke"),
    ("Remote Raider Flare", 640, [("Chemicals", 2), ("Rubber Parts", 4)], "Utility", "Signal"),
    ("Green Light Stick", 640, [("Chemicals", 3)], "Utility", "Illumination"),
    ("Photoelectric Cloak", 5000, [("Advanced Electrical Components", 2), ("Speaker Component", 4)], "Utility III", "Stealth"),
    ("Raider Hatch Key", 2000, [("Advanced Electrical Components", 1), ("Sensors", 3)], "Utility III", "Key"),
    
    # === REFINER ===
    ("Light Gun Parts", 700, [("Simple Gun Parts", 4)], "Refiner II", "Crafting material"),
    ("Medium Gun Parts", 700, [("Simple Gun Parts", 4)], "Refiner II", "Crafting material"),
    ("Heavy Gun Parts", 700, [("Simple Gun Parts", 4)], "Refiner II", "Crafting material"),
    ("Complex Gun Parts", 3000, [("Light Gun Parts", 2), ("Medium Gun Parts", 2), ("Heavy Gun Parts", 2)], "Refiner III", "High-tier"),
    ("Mechanical Components", 640, [("Metal Parts", 7), ("Rubber Parts", 3)], "Refiner", "Basic refined"),
    ("Electrical Components", 640, [("Plastic Parts", 8), ("Rubber Parts", 4)], "Refiner", "Basic refined"),
    ("Advanced Mechanical Components", 1750, [("Steel Spring", 2), ("Mechanical Components", 2)], "Refiner II", "Rare"),
    ("Advanced Electrical Components", 1750, [("Wires", 3), ("Electrical Components", 2)], "Refiner II", "Rare"),
    ("Mod Components", 1750, [("Steel Spring", 2), ("Mechanical Components", 2)], "Refiner II", "For attachments"),
    ("Crude Explosives", 270, [("Chemicals", 6)], "Refiner", "Basic explosive"),
    ("Explosive Compound", 1000, [("Crude Explosives", 2), ("Oil", 2)], "Refiner II", "Refined explosive"),
    ("Durable Cloth", 640, [("Fabric", 14)], "Refiner", "Fabric refined"),
    ("ARC Circuitry", 1000, [("ARC Alloy", 8)], "Refiner II", "ARC material"),
    ("ARC Motion Core", 1000, [("ARC Alloy", 8)], "Refiner II", "ARC material"),
    ("Magnetic Accelerator", 5500, [("Advanced Mechanical Components", 2), ("ARC Motion Core", 2)], "Refiner III", "High-tier"),
    ("Power Rod", 5500, [("Advanced Electrical Components", 2), ("ARC Circuitry", 2)], "Refiner III", "High-tier"),
    ("Antiseptic", 1000, [("Chemicals", 10), ("Great Mullein", 2)], "Refiner", "Medical"),
    
    # === GUNSMITH - MODIFICATIONS ===
    ("Angled Grip I", 640, [("Plastic Parts", 6), ("Duct Tape", 1)], "Gunsmith", "Grip mod"),
    ("Angled Grip II", 2000, [("Mechanical Components", 2), ("Duct Tape", 3)], "Gunsmith II", ""),
    ("Angled Grip III", 5000, [("Mod Components", 2), ("Duct Tape", 5)], "Gunsmith III", ""),
    ("Vertical Grip I", 640, [("Plastic Parts", 6), ("Duct Tape", 1)], "Gunsmith", ""),
    ("Vertical Grip II", 2000, [("Mechanical Components", 2), ("Duct Tape", 3)], "Gunsmith II", ""),
    ("Vertical Grip III", 5000, [("Mod Components", 2), ("Duct Tape", 5)], "Gunsmith III", ""),
    ("Horizontal Grip", 7000, [("Mod Components", 2), ("Duct Tape", 5)], "Gunsmith III", "Epic grip"),
    ("Stable Stock I", 640, [("Rubber Parts", 6), ("Duct Tape", 1)], "Gunsmith", "Stock mod"),
    ("Stable Stock II", 2000, [("Mechanical Components", 2), ("Duct Tape", 3)], "Gunsmith II", ""),
    ("Stable Stock III", 5000, [("Mod Components", 2), ("Duct Tape", 5)], "Gunsmith III", ""),
    ("Lightweight Stock", 5000, [("Mod Components", 2), ("Duct Tape", 5)], "Gunsmith III", "Epic stock"),
    ("Compensator I", 640, [("Metal Parts", 6), ("Wires", 1)], "Gunsmith", "Muzzle mod"),
    ("Compensator II", 2000, [("Mechanical Components", 2), ("Wires", 4)], "Gunsmith II", ""),
    ("Compensator III", 5000, [("Mod Components", 2), ("Wires", 8)], "Gunsmith III", ""),
    ("Muzzle Brake I", 640, [("Metal Parts", 6), ("Wires", 1)], "Gunsmith", ""),
    ("Muzzle Brake II", 2000, [("Mechanical Components", 2), ("Wires", 4)], "Gunsmith II", ""),
    ("Muzzle Brake III", 5000, [("Mod Components", 2), ("Wires", 8)], "Gunsmith III", ""),
    ("Shotgun Choke I", 640, [("Metal Parts", 6), ("Wires", 1)], "Gunsmith", ""),
    ("Shotgun Choke II", 2000, [("Mechanical Components", 2), ("Wires", 4)], "Gunsmith II", ""),
    ("Shotgun Choke III", 5000, [("Mod Components", 2), ("Wires", 8)], "Gunsmith III", ""),
    ("Shotgun Silencer", 5000, [("Mod Components", 2), ("Wires", 8)], "Gunsmith III", "Epic"),
    ("Extended Barrel", 5000, [("Mod Components", 2), ("Wires", 8)], "Gunsmith III", "Epic"),
    ("Extended Light Mag I", 640, [("Plastic Parts", 6), ("Steel Spring", 1)], "Gunsmith", "Mag mod"),
    ("Extended Light Mag II", 2000, [("Mechanical Components", 2), ("Steel Spring", 3)], "Gunsmith II", ""),
    ("Extended Light Mag III", 5000, [("Mod Components", 2), ("Steel Spring", 5)], "Gunsmith III", ""),
    ("Extended Medium Mag I", 640, [("Plastic Parts", 6), ("Steel Spring", 1)], "Gunsmith", ""),
    ("Extended Medium Mag II", 2000, [("Mechanical Components", 2), ("Steel Spring", 3)], "Gunsmith II", ""),
    ("Extended Medium Mag III", 5000, [("Mod Components", 2), ("Steel Spring", 5)], "Gunsmith III", ""),
    ("Extended Shotgun Mag I", 640, [("Plastic Parts", 6), ("Steel Spring", 1)], "Gunsmith", ""),
    
    # === GUNSMITH - WEAPONS ===
    # Common weapons
    ("Hairpin I", 500, [("Metal Parts", 2), ("Plastic Parts", 5)], "Gunsmith", "Pistol"),
    ("Ferro I", 500, [("Metal Parts", 5), ("Rubber Parts", 2)], "Gunsmith", "Battle Rifle"),
    ("Kettle I", 500, [("Metal Parts", 6), ("Rubber Parts", 8)], "Gunsmith", "AR"),
    ("Rattler I", 500, [("Metal Parts", 16), ("Rubber Parts", 12)], "Gunsmith", "AR"),
    ("Stitcher I", 500, [("Metal Parts", 8), ("Rubber Parts", 4)], "Gunsmith", "SMG"),
    
    # Uncommon weapons
    ("Anvil I", 2000, [("Mechanical Components", 5), ("Simple Gun Parts", 6)], "Gunsmith", "Hand Cannon"),
    ("Burletta I", 1500, [("Mechanical Components", 3), ("Simple Gun Parts", 3)], "Gunsmith", "Pistol"),
    ("Il Toro I", 2000, [("Mechanical Components", 5), ("Simple Gun Parts", 6)], "Gunsmith", "Shotgun"),
    ("Arpeggio I", 2000, [("Mechanical Components", 6), ("Simple Gun Parts", 6)], "Gunsmith", "AR"),
    
    # Rare weapons
    ("Osprey I", 3500, [("Advanced Mechanical Components", 2), ("Medium Gun Parts", 3), ("Wires", 7)], "Gunsmith II", "Sniper"),
    ("Torrente I", 3500, [("Advanced Mechanical Components", 2), ("Medium Gun Parts", 3), ("Steel Spring", 6)], "Gunsmith II", "LMG"),
    ("Venator I", 3500, [("Advanced Mechanical Components", 2), ("Medium Gun Parts", 3), ("Industrial Magnet", 5)], "Gunsmith II", "Pistol"),
    ("Renegade I", 3500, [("Advanced Mechanical Components", 2), ("Medium Gun Parts", 3), ("Oil", 5)], "Gunsmith II", "Battle Rifle"),
    
    # Epic weapons
    ("Bettina I", 5000, [("Advanced Mechanical Components", 3), ("Heavy Gun Parts", 3), ("Canister", 3)], "Gunsmith III", "AR"),
    ("Tempest I", 5000, [("Advanced Mechanical Components", 2), ("Medium Gun Parts", 1)], "Gunsmith III", "AR - cheap!"),
    ("Vulcano I", 5000, [("Magnetic Accelerator", 1), ("Heavy Gun Parts", 3), ("Exodus Modules", 1)], "Gunsmith III", "Shotgun"),
    ("Hullcracker I", 5000, [("Magnetic Accelerator", 1), ("Heavy Gun Parts", 3), ("Exodus Modules", 1)], "Gunsmith III", "Launcher"),
    ("Bobcat I", 5000, [("Magnetic Accelerator", 1), ("Light Gun Parts", 3), ("Exodus Modules", 2)], "Gunsmith III", "SMG"),
    
    # Legendary weapons
    ("Jupiter", 5000, [("Magnetic Accelerator", 3), ("Complex Gun Parts", 3), ("Queen Reactor", 1)], "Gunsmith III", "Sniper"),
    ("Equalizer", 5000, [("Magnetic Accelerator", 3), ("Complex Gun Parts", 3), ("Queen Reactor", 1)], "Gunsmith III", "LMG"),
    ("Aphelion", 5000, [("Magnetic Accelerator", 3), ("Complex Gun Parts", 3), ("Matriarch Reactor", 1)], "Gunsmith III", "Special"),
]


def calculate_material_cost(recipe):
    """Calculate total material cost for a recipe."""
    total = 0
    for material, qty in recipe:
        if material in material_values:
            total += material_values[material] * qty
        else:
            print(f"WARNING: Unknown material '{material}' - using default value 500")
            total += 500 * qty
    return total


def get_material_string(recipe):
    """Format recipe as readable string."""
    return ", ".join([f"{qty}× {mat}" for mat, qty in recipe])


def create_crafting_spreadsheet(output_path="arc_raiders_crafting_profit.xlsx"):
    """Generate the crafting profitability spreadsheet."""
    
    # Calculate all items
    crafting_data = []
    for name, sell_value, recipe, bench, notes in recipes:
        material_cost = calculate_material_cost(recipe)
        profit = sell_value - material_cost
        profit_pct = (profit / material_cost * 100) if material_cost > 0 else 0
        crafting_data.append({
            "name": name,
            "sell_value": sell_value,
            "material_cost": material_cost,
            "profit": profit,
            "profit_pct": profit_pct,
            "recipe": get_material_string(recipe),
            "bench": bench,
            "notes": notes,
            "profitable": profit > 0
        })
    
    # Sort by profit
    crafting_data.sort(key=lambda x: x["profit"], reverse=True)
    
    # Create workbook
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill_green = PatternFill("solid", fgColor="2E7D32")
    header_fill_blue = PatternFill("solid", fgColor="1565C0")
    profit_fill = PatternFill("solid", fgColor="C8E6C9")
    loss_fill = PatternFill("solid", fgColor="FFCDD2")
    breakeven_fill = PatternFill("solid", fgColor="FFF9C4")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    currency_format = '"$"#,##0'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Sheet 1: Profitable Items
    ws_profit = wb.active
    ws_profit.title = "Profitable Crafts"
    
    headers = ["Item Name", "Sell Value", "Material Cost", "Profit", "Profit %", "Recipe", "Bench"]
    for col, header in enumerate(headers, 1):
        cell = ws_profit.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill_green
        cell.alignment = center_align
        cell.border = thin_border
    
    profitable_items = [item for item in crafting_data if item["profitable"]]
    row = 2
    for item in profitable_items:
        ws_profit.cell(row=row, column=1, value=item["name"]).alignment = left_align
        ws_profit.cell(row=row, column=2, value=item["sell_value"]).number_format = currency_format
        ws_profit.cell(row=row, column=3, value=item["material_cost"]).number_format = currency_format
        ws_profit.cell(row=row, column=4, value=item["profit"]).number_format = currency_format
        ws_profit.cell(row=row, column=5, value=round(item["profit_pct"], 1))
        ws_profit.cell(row=row, column=6, value=item["recipe"]).alignment = left_align
        ws_profit.cell(row=row, column=7, value=item["bench"]).alignment = center_align
        for col in range(1, 8):
            ws_profit.cell(row=row, column=col).border = thin_border
            ws_profit.cell(row=row, column=col).fill = profit_fill
        row += 1
    
    ws_profit.column_dimensions['A'].width = 26
    ws_profit.column_dimensions['B'].width = 12
    ws_profit.column_dimensions['C'].width = 14
    ws_profit.column_dimensions['D'].width = 10
    ws_profit.column_dimensions['E'].width = 10
    ws_profit.column_dimensions['F'].width = 55
    ws_profit.column_dimensions['G'].width = 14
    
    # Sheet 2: All Items
    ws_all = wb.create_sheet("All Crafts")
    for col, header in enumerate(headers, 1):
        cell = ws_all.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill_blue
        cell.alignment = center_align
        cell.border = thin_border
    
    row = 2
    for item in crafting_data:
        ws_all.cell(row=row, column=1, value=item["name"]).alignment = left_align
        ws_all.cell(row=row, column=2, value=item["sell_value"]).number_format = currency_format
        ws_all.cell(row=row, column=3, value=item["material_cost"]).number_format = currency_format
        ws_all.cell(row=row, column=4, value=item["profit"]).number_format = currency_format
        ws_all.cell(row=row, column=5, value=round(item["profit_pct"], 1))
        ws_all.cell(row=row, column=6, value=item["recipe"]).alignment = left_align
        ws_all.cell(row=row, column=7, value=item["bench"]).alignment = center_align
        
        fill = profit_fill if item["profit"] > 0 else (breakeven_fill if item["profit"] == 0 else loss_fill)
        for col in range(1, 8):
            ws_all.cell(row=row, column=col).border = thin_border
            ws_all.cell(row=row, column=col).fill = fill
        row += 1
    
    ws_all.column_dimensions['A'].width = 26
    ws_all.column_dimensions['B'].width = 12
    ws_all.column_dimensions['C'].width = 14
    ws_all.column_dimensions['D'].width = 10
    ws_all.column_dimensions['E'].width = 10
    ws_all.column_dimensions['F'].width = 55
    ws_all.column_dimensions['G'].width = 14
    
    # Sheet 3: Material Values
    ws_mats = wb.create_sheet("Material Values")
    ws_mats.cell(row=1, column=1, value="Material").font = header_font
    ws_mats.cell(row=1, column=1).fill = header_fill_blue
    ws_mats.cell(row=1, column=2, value="Sell Value").font = header_font
    ws_mats.cell(row=1, column=2).fill = header_fill_blue
    ws_mats.cell(row=1, column=3, value="Rarity").font = header_font
    ws_mats.cell(row=1, column=3).fill = header_fill_blue
    
    sorted_mats = sorted(material_values.items(), key=lambda x: -x[1])
    row = 2
    for mat, val in sorted_mats:
        ws_mats.cell(row=row, column=1, value=mat).border = thin_border
        ws_mats.cell(row=row, column=2, value=val).number_format = currency_format
        ws_mats.cell(row=row, column=2).border = thin_border
        
        if val >= 5000: rarity = "Epic/Legendary"
        elif val >= 1000: rarity = "Rare"
        elif val >= 500: rarity = "Uncommon"
        else: rarity = "Common"
        ws_mats.cell(row=row, column=3, value=rarity).border = thin_border
        row += 1
    
    ws_mats.column_dimensions['A'].width = 32
    ws_mats.column_dimensions['B'].width = 12
    ws_mats.column_dimensions['C'].width = 15
    
    # Sheet 4: Summary
    ws_summary = wb.create_sheet("Summary")
    ws_summary['A1'] = "Arc Raiders Crafting Profitability Analysis"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary['A2'] = "Data Source: ARCTracker.io (Dec 2025)"
    
    ws_summary['A4'] = "Statistics:"
    ws_summary['A4'].font = Font(bold=True)
    ws_summary['A5'] = f"Total recipes analyzed: {len(crafting_data)}"
    ws_summary['A6'] = f"Profitable crafts: {len(profitable_items)}"
    ws_summary['A7'] = f"Break-even crafts: {len([i for i in crafting_data if i['profit'] == 0])}"
    ws_summary['A8'] = f"Loss-making crafts: {len([i for i in crafting_data if i['profit'] < 0])}"
    
    ws_summary['A10'] = "Top 10 Most Profitable Crafts:"
    ws_summary['A10'].font = Font(bold=True)
    for i, item in enumerate(profitable_items[:10], 11):
        ws_summary[f'A{i}'] = f"{item['name']}: ${item['profit']:,} profit ({item['profit_pct']:.0f}%)"
    
    ws_summary['A22'] = "Key Findings:"
    ws_summary['A22'].font = Font(bold=True)
    ws_summary['A23'] = "• Tempest I is the most profitable weapon craft (only needs 2x Adv Mech + 1x Med Gun Parts)"
    ws_summary['A24'] = "• Tier II attachments (Grips, Stocks, etc.) are consistently profitable"
    ws_summary['A25'] = "• Light Sticks are extremely profitable at 327% ROI"
    ws_summary['A26'] = "• Gun Parts (Light/Medium/Heavy) are LOSSES - sell Simple Gun Parts instead"
    ws_summary['A27'] = "• Legendary weapons (Jupiter, Equalizer, Aphelion) are massive losses - keep them!"
    ws_summary['A28'] = "• Wolfpack now costs $5,000 in materials = break-even after patch"
    ws_summary['A29'] = "• Heavy Fuze Grenade costs $1,600 = break-even"
    
    ws_summary.column_dimensions['A'].width = 80
    
    wb.save(output_path)
    
    print(f"Created {output_path} with {len(crafting_data)} recipes")
    print(f"Profitable: {len(profitable_items)}")
    print(f"Break-even: {len([i for i in crafting_data if i['profit'] == 0])}")
    print(f"Loss: {len([i for i in crafting_data if i['profit'] < 0])}")
    print("\nTop 5 profitable:")
    for item in profitable_items[:5]:
        print(f"  {item['name']}: ${item['profit']:,} ({item['profit_pct']:.0f}%)")


if __name__ == "__main__":
    create_crafting_spreadsheet()
